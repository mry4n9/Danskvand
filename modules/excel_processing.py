import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils.dataframe import dataframe_to_rows # Not needed for cell-by-cell writing
from openpyxl.utils import get_column_letter
import io

# --- Styling Functions (Unchanged) ---
def apply_header_style(cell):
    """Applies style to header cells (typically in Column A)."""
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) # Wrap text for headers too
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def apply_content_style(cell):
    """Applies style to content cells (data columns)."""
    cell.alignment = Alignment(vertical="center", wrap_text=True) # Top align for readability
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def adjust_column_width_and_row_height(worksheet):
    """Adjusts column widths and sets a minimum row height."""
    # Adjust Column A width based on headers
    max_length_a = 0
    for cell in worksheet['A']:
         if cell.value:
            max_length_a = max(max_length_a, len(str(cell.value)))
    worksheet.column_dimensions['A'].width = min((max_length_a + 2) * 1.2, 30) # Max width 30 for headers

    # Adjust other columns based on content
    for col_idx in range(2, worksheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                # Estimate length considering potential line breaks in wrapped text
                cell_lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in cell_lines) if cell_lines else 0
                max_length = max(max_length, max_line_length)

        adjusted_width = min((max_length + 2) * 1.2, 50) # Max width 50
        # Set a minimum width as well
        worksheet.column_dimensions[column_letter].width = max(adjusted_width, 15)

    # Set a minimum row height for wrapped text visibility
    for row_idx in range(1, worksheet.max_row + 1):
         # Simple approach: Set a default height or slightly increase based on content length
         # A more complex approach would estimate height based on text length and column width
         worksheet.row_dimensions[row_idx].height = 30 # Default height, adjust as needed


def create_excel_report(ad_data: dict, company_name: str, lead_objective: str) -> bytes:
    """
    Creates an XLSX report with headers in Column A and data in subsequent columns.
    ad_data is a dictionary where keys are like "Email", "LinkedIn_BA", "GoogleSearch"
    and values are lists of ad dicts or a single dict for Google Ads.
    """
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    # --- Email Sheet ---
    if "Email" in ad_data and ad_data["Email"] and ad_data["Email"].get("emails"):
        ws_email = wb.create_sheet("Email")
        email_headers = ["Ad Name", "Funnel Stage", "Headline", "Subject Line", "Body", "CTA"]
        
        # Write headers to Column A
        for i, header in enumerate(email_headers, start=1):
            cell = ws_email.cell(row=i, column=1, value=header)
            apply_header_style(cell)

        # Write ad data to columns B, C, ...
        emails = ad_data["Email"].get("emails", [])
        for j, ad in enumerate(emails, start=2): # Start data from Column B (index 2)
            ad_name = f"Email_Demand Capture_Ver. {j-1}"
            data_points = [
                ad_name,
                "Demand Capture",
                ad.get("headline", ""),
                ad.get("subject_line", ""),
                ad.get("body", ""),
                ad.get("cta", "")
            ]
            # Write data vertically in the current column
            for i, value in enumerate(data_points, start=1):
                cell = ws_email.cell(row=i, column=j, value=value)
                apply_content_style(cell)
        
        adjust_column_width_and_row_height(ws_email)

    # --- LinkedIn Sheet ---
    linkedin_ads_present = any(k.startswith("LinkedIn") and ad_data[k] and ad_data[k].get(f"linkedin_{k.split('_')[1].lower()}_ads") for k in ad_data)
    if linkedin_ads_present:
        ws_linkedin = wb.create_sheet("LinkedIn")
        linkedin_headers = ["Ad Name", "Funnel Stage", "Introductory Text", "Image Copy", "Headline", "Destination", "CTA Button"]
        
        # Write headers to Column A
        for i, header in enumerate(linkedin_headers, start=1):
            cell = ws_linkedin.cell(row=i, column=1, value=header)
            apply_header_style(cell)

        current_col_index = 2 # Start data from Column B
        
        for funnel_key, funnel_stage_name, ads_list_key in [
            ("LinkedIn_BA", "Brand Awareness", "linkedin_brand_awareness_ads"),
            ("LinkedIn_DG", "Demand Gen", "linkedin_demand_gen_ads"),
            ("LinkedIn_DC", "Demand Capture", "linkedin_demand_capture_ads")
        ]:
            if funnel_key in ad_data and ad_data[funnel_key] and ad_data[funnel_key].get(ads_list_key):
                ads = ad_data[funnel_key].get(ads_list_key, [])
                for ad_index, ad in enumerate(ads):
                    ad_name = f"LinkedIn_{funnel_stage_name.replace(' ', '')}_Ver. {ad_index+1}"
                    data_points = [
                        ad_name,
                        funnel_stage_name,
                        ad.get("introductory_text", ""),
                        ad.get("image_copy", ""),
                        ad.get("headline", ""),
                        ad.get("destination_url", ""),
                        ad.get("cta_button", "")
                    ]
                    # Write data vertically in the current column
                    for i, value in enumerate(data_points, start=1):
                        cell = ws_linkedin.cell(row=i, column=current_col_index, value=value)
                        apply_content_style(cell)
                    current_col_index += 1 # Move to the next column for the next ad

        adjust_column_width_and_row_height(ws_linkedin)

    # --- Facebook Sheet ---
    facebook_ads_present = any(k.startswith("Facebook") and ad_data[k] and ad_data[k].get(f"facebook_{k.split('_')[1].lower()}_ads") for k in ad_data)
    if facebook_ads_present:
        ws_facebook = wb.create_sheet("FaceBook") # Note: 'FaceBook' as per spec
        facebook_headers = ["Ad Name", "Funnel Stage", "Primary Text", "Image Copy", "Headline", "Link Description", "Destination", "CTA Button"]
        
        # Write headers to Column A
        for i, header in enumerate(facebook_headers, start=1):
            cell = ws_facebook.cell(row=i, column=1, value=header)
            apply_header_style(cell)

        current_col_index = 2 # Start data from Column B

        for funnel_key, funnel_stage_name, ads_list_key in [
            ("Facebook_BA", "Brand Awareness", "facebook_brand_awareness_ads"),
            ("Facebook_DG", "Demand Gen", "facebook_demand_gen_ads"),
            ("Facebook_DC", "Demand Capture", "facebook_demand_capture_ads")
        ]:
             if funnel_key in ad_data and ad_data[funnel_key] and ad_data[funnel_key].get(ads_list_key):
                ads = ad_data[funnel_key].get(ads_list_key, [])
                for ad_index, ad in enumerate(ads):
                    ad_name = f"Facebook_{funnel_stage_name.replace(' ', '')}_Ver. {ad_index+1}"
                    data_points = [
                        ad_name,
                        funnel_stage_name,
                        ad.get("primary_text", ""),
                        ad.get("image_copy", ""),
                        ad.get("headline", ""),
                        ad.get("link_description", ""),
                        ad.get("destination_url", ""),
                        ad.get("cta_button", "")
                    ]
                    # Write data vertically in the current column
                    for i, value in enumerate(data_points, start=1):
                        cell = ws_facebook.cell(row=i, column=current_col_index, value=value)
                        apply_content_style(cell)
                    current_col_index += 1 # Move to the next column for the next ad

        adjust_column_width_and_row_height(ws_facebook)

    # --- Google Search Sheet ---
    if "GoogleSearch" in ad_data and ad_data["GoogleSearch"]:
        ws_gsearch = wb.create_sheet("Google Search")
        gsearch_headers = ["Headline", "Description"]
        
        # Write headers to Column A
        for i, header in enumerate(gsearch_headers, start=1):
            cell = ws_gsearch.cell(row=i, column=1, value=header)
            apply_header_style(cell)

        headlines = ad_data["GoogleSearch"].get("headlines", [])
        descriptions = ad_data["GoogleSearch"].get("descriptions", [])
        
        # Write headlines horizontally starting from B1
        for j, headline in enumerate(headlines, start=2):
            cell = ws_gsearch.cell(row=1, column=j, value=headline)
            apply_content_style(cell)
            
        # Write descriptions horizontally starting from B2
        for j, description in enumerate(descriptions, start=2):
            cell = ws_gsearch.cell(row=2, column=j, value=description)
            apply_content_style(cell)

        adjust_column_width_and_row_height(ws_gsearch)

    # --- Google Display Sheet ---
    if "GoogleDisplay" in ad_data and ad_data["GoogleDisplay"]:
        ws_gdisplay = wb.create_sheet("Google Display")
        gdisplay_headers = ["Headline", "Description"]
        
        # Write headers to Column A
        for i, header in enumerate(gdisplay_headers, start=1):
            cell = ws_gdisplay.cell(row=i, column=1, value=header)
            apply_header_style(cell)

        headlines = ad_data["GoogleDisplay"].get("headlines", [])
        descriptions = ad_data["GoogleDisplay"].get("descriptions", [])

        # Write headlines horizontally starting from B1
        for j, headline in enumerate(headlines, start=2):
            cell = ws_gdisplay.cell(row=1, column=j, value=headline)
            apply_content_style(cell)
            
        # Write descriptions horizontally starting from B2
        for j, description in enumerate(descriptions, start=2):
            cell = ws_gdisplay.cell(row=2, column=j, value=description)
            apply_content_style(cell)

        adjust_column_width_and_row_height(ws_gdisplay)

    # Save to a BytesIO object
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes.getvalue()